"""
server/routers/history_archives.py — browse saved history CSV files.

GET  /api/v1/history/archives               — list all CSV files in logs/
GET  /api/v1/history/archives/{filename}    — parse and return entries
GET  /api/v1/history/archives/{filename}/xlsx — download full Excel report (3 sheets)
GET  /api/v1/history/archives/{filename}/raw — download raw CSV file
DEL  /api/v1/history/archives/{filename}    — delete archive
"""

from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import FileResponse, StreamingResponse

from server.routers.auth import get_current_user

router = APIRouter(prefix="/api/v1/history", tags=["history-archives"])

_SESSION_ID_RE = re.compile(r"_([0-9a-f]{6,})\.csv$", re.IGNORECASE)


def _session_id_from_csv(filename: str) -> str | None:
    m = _SESSION_ID_RE.search(filename)
    return m.group(1) if m else None


def _build_xlsx(log_dir: Path, csv_path: Path) -> io.BytesIO:
    """Generate an Excel workbook with Commands, Artifacts and Uploads sheets."""
    session_id = _session_id_from_csv(csv_path.name)

    cmd_rows: list[list] = []
    art_rows: list[list] = []
    # Track artifacts: added vs removed
    art_state: dict = {}

    try:
        with open(csv_path, newline="", encoding="utf-8", errors="replace") as f:
            for row in csv.reader(f):
                if len(row) < 4 or row[0].startswith("---") or row[0] == "timestamp":
                    continue
                ts  = row[0]
                sid = row[1] if len(row) > 1 else ""
                cid = row[2] if len(row) > 2 else ""
                cmd = row[3] if len(row) > 3 else ""
                resp     = row[4] if len(row) > 4 else ""
                operator = row[5] if len(row) > 5 else ""

                if cid == "ARTIFACT":
                    kind, _, path = cmd.partition(":")
                    if kind and path:
                        art_state[(kind, path)] = {"type": kind, "path": path, "recorded_at": ts, "removed": False}
                elif cid == "ARTIFACT_REMOVED":
                    kind, _, path = cmd.partition(":")
                    if (kind, path) in art_state:
                        art_state[(kind, path)]["removed"] = True
                else:
                    cmd_rows.append([ts, sid, cid, operator, cmd, resp])
    except Exception:
        pass

    # Uploads
    ul_rows: list[list] = []
    if session_id:
        ul_path = log_dir / f"uploads_{session_id}.json"
        if ul_path.exists():
            try:
                records = json.loads(ul_path.read_text())
                for r in records:
                    ul_rows.append([
                        r.get("timestamp", ""),
                        r.get("filename", ""),
                        r.get("remote_path", ""),
                        r.get("size", ""),
                        r.get("cmd_id", ""),
                        r.get("status", "on target"),
                    ])
            except Exception:
                pass

    wb = openpyxl.Workbook()

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1F1F2E")
    hdr_align = Alignment(horizontal="center")

    _FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")

    def _safe(v: object) -> object:
        if isinstance(v, str) and v.startswith(_FORMULA_PREFIXES):
            return "'" + v
        return v

    def _add_sheet(title: str, headers: list, rows: list):
        ws = wb.create_sheet(title=title)
        ws.append(headers)
        for cell in ws[1]:
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.alignment = hdr_align
        for row in rows:
            ws.append([_safe(v) for v in row])
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    _add_sheet(
        "Commands",
        ["Timestamp", "Session ID", "CMD ID", "Operator", "Command", "Response"],
        cmd_rows,
    )
    _add_sheet(
        "Artifacts",
        ["Timestamp", "Type", "Path", "Status"],
        [
            [a["recorded_at"], a["type"], a["path"], "removed" if a["removed"] else "on target"]
            for a in art_state.values()
        ],
    )
    _add_sheet(
        "Uploads",
        ["Timestamp", "Filename", "Remote Path", "Size (bytes)", "CMD ID", "Status"],
        ul_rows,
    )

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/archives")
def list_archives(request: Request, _: str = Depends(get_current_user)):
    log_dir = Path(request.app.state.cfg.log_dir)
    files = []
    for f in sorted(log_dir.glob("*.csv"), key=lambda x: x.stat().st_mtime, reverse=True):
        stat = f.stat()
        files.append({
            "filename":    f.name,
            "size_bytes":  stat.st_size,
            "modified_at": stat.st_mtime,
        })
    return files


@router.get("/archives/{filename}/artifacts")
def read_archive_artifacts(filename: str, request: Request, _: str = Depends(get_current_user)):
    """Return on-target artifacts and uploads for a CSV archive."""
    if "/" in filename or "\\" in filename or ".." in filename or "\x00" in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    log_dir  = Path(request.app.state.cfg.log_dir)
    csv_path = log_dir / filename
    if not csv_path.exists() or csv_path.suffix != ".csv":
        raise HTTPException(status_code=404, detail="Archive not found")

    session_id = _session_id_from_csv(filename)
    art_state: dict = {}

    try:
        with open(csv_path, newline="", encoding="utf-8", errors="replace") as f:
            for row in csv.reader(f):
                if len(row) < 4 or row[0].startswith("---") or row[0] == "timestamp":
                    continue
                ts, cid, cmd = row[0], row[2], row[3]
                if cid == "ARTIFACT":
                    kind, _, path = cmd.partition(":")
                    if kind and path:
                        art_state[(kind, path)] = {"type": kind, "path": path, "recorded_at": ts, "status": "on target"}
                elif cid == "ARTIFACT_REMOVED":
                    kind, _, path = cmd.partition(":")
                    if (kind, path) in art_state:
                        art_state[(kind, path)]["status"] = "removed"
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to parse archive")

    uploads: list = []
    if session_id:
        ul_path = log_dir / f"uploads_{session_id}.json"
        if ul_path.exists():
            try:
                uploads = json.loads(ul_path.read_text())
            except Exception:
                uploads = []

    return {"on_target": list(art_state.values()), "uploads": uploads}


@router.get("/archives/{filename}/xlsx")
def download_archive_xlsx(filename: str, request: Request, _: str = Depends(get_current_user)):
    """Generate and serve a full Excel report (Commands + Artifacts + Uploads)."""
    if "/" in filename or "\\" in filename or ".." in filename or "\x00" in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    log_dir = Path(request.app.state.cfg.log_dir)
    csv_path = log_dir / filename
    if not csv_path.exists() or csv_path.suffix != ".csv":
        raise HTTPException(status_code=404, detail="Archive not found")

    buf = _build_xlsx(log_dir, csv_path)
    xlsx_name = csv_path.stem + ".xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{xlsx_name}"'},
    )


@router.get("/archives/{filename}")
def read_archive(filename: str, request: Request, _: str = Depends(get_current_user)):
    if "/" in filename or "\\" in filename or ".." in filename or "\x00" in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    log_dir = Path(request.app.state.cfg.log_dir)
    path    = log_dir / filename
    if not path.exists() or path.suffix != ".csv":
        raise HTTPException(status_code=404, detail="Archive not found")

    by_cmd_id:     dict = {}
    no_id_entries: list = []
    try:
        with open(path, newline="") as f:
            for row in csv.reader(f):
                if len(row) < 5 or row[0].startswith("---") or row[0] == "timestamp":
                    continue
                ts, sid, cid, cmd, resp = row[0], row[1], row[2], row[3], row[4]
                operator = row[5] if len(row) > 5 else ""
                if cid in ("ARTIFACT", "ARTIFACT_REMOVED"):
                    continue
                entry = {
                    "timestamp":  ts,
                    "session_id": sid,
                    "cmd_id":     cid,
                    "command":    cmd,
                    "response":   resp,
                    "operator":   operator,
                }
                if cid:
                    by_cmd_id[cid] = entry
                else:
                    no_id_entries.append(entry)
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to parse archive")

    entries = list(by_cmd_id.values()) + no_id_entries
    entries.sort(key=lambda e: e["timestamp"])
    return entries


@router.delete("/archives/{filename}", status_code=204)
async def delete_archive(filename: str, request: Request, username: str = Depends(get_current_user)):
    if "/" in filename or "\\" in filename or ".." in filename or "\x00" in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    log_dir = Path(request.app.state.cfg.log_dir)
    path    = log_dir / filename
    if not path.exists() or path.suffix != ".csv":
        raise HTTPException(status_code=404, detail="Archive not found")
    path.unlink()
    ws = request.app.state.ws
    await ws.broadcast({"type": "history.archive.deleted", "payload": {"filename": filename, "deleted_by": username}})


@router.get("/archives/{filename}/raw")
def download_archive_raw(filename: str, request: Request, _: str = Depends(get_current_user)):
    """Serve the raw CSV file as a download."""
    if "/" in filename or "\\" in filename or ".." in filename or "\x00" in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    log_dir = Path(request.app.state.cfg.log_dir)
    path    = log_dir / filename
    if not path.exists() or path.suffix != ".csv":
        raise HTTPException(status_code=404, detail="Archive not found")
    return FileResponse(
        path=str(path),
        media_type="text/csv",
        filename=filename,
    )
